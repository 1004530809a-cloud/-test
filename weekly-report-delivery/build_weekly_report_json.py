import argparse
import json
import re
import zipfile
from copy import copy
from pathlib import Path
from xml.etree import ElementTree as ET

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


SZA_KEYWORD = "华为"
HEADER_REQUIRED = {
    "advertiser": "广告主名称",
    "purchase_order_no": "采购订单编号",
    "exec_price_tax": "执行价(含税)",
    "gross_profit": "预估订单毛利额",
    "gross_margin": "预估订单毛利率",
    "actual_income_no_tax": "订单实际收入(去税)",
    "price_result": "压价结果",
    "rebate_tax_third": "应约时返点金额含税(三方)",
    "rebate_tax_non_third": "应约时返点金额含税(非三方)",
    "rebate_tax_platform": "应约时返点金额含税(平台)",
    "rebate_no_tax_estimate": "预估返点额不含税",
    "spu_category": "SPU类目",
    "owner_media": "行业媒介",
}
EXCLUDED_SPU_KEYWORD = "汽车"
LOW_MARGIN_THRESHOLD = 14
YELLOW_FILL_RGB = "FFFFFF00"
LOW_MARGIN_REASONS = [
    "sza年框账号",
    "客户指定合作账号",
    "二压已最高，无法提升",
    "对客折扣高",
    "koc批量打包订单",
    "特殊订单",
    "机构年框政策，无法再提升",
    "冲榜机构固定政策",
    "非行业媒介压价",
    "未压价",
    "客户走单账号",
]
DEFAULT_FEEDBACK_TEMPLATE = "【0912更新】行业二组-周报模板.xlsx"
DEFAULT_FEEDBACK_TEMPLATE_SHEET = "3月影响毛利订单反馈"


def fmt_money(value: float) -> str:
    return f"{value:,.2f}"


def fmt_percent(numerator: float, denominator: float) -> str:
    if not denominator:
        return "0.00%"
    return f"{(numerator / denominator) * 100:.2f}%"


def to_float(value) -> float:
    if value in (None, "", "-"):
        return 0.0
    try:
        return float(value)
    except Exception:
        text = str(value).replace(",", "").strip()
        try:
            return float(text)
        except Exception:
            return 0.0


def to_percent_number(value) -> float:
    if value in (None, "", "-"):
        return 0.0
    raw_text = str(value).strip()
    has_percent_sign = "%" in raw_text
    text = raw_text.replace("%", "").replace(",", "").strip()
    try:
        parsed = float(text)
    except Exception:
        return 0.0
    if has_percent_sign:
        return parsed
    return parsed * 100 if abs(parsed) <= 1 else parsed


def to_file_href(path: Path) -> str:
    return path.resolve().as_uri()


def extract_owner_name(value) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    parts = []
    for chunk in text.strip("[]").split(","):
        piece = chunk.strip()
        if not piece:
            continue
        name = piece.split("/")[-1].strip()
        if name:
            parts.append(name)
    return "、".join(dict.fromkeys(parts))


def find_header_row(ws) -> tuple[int, list[str]]:
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 10), values_only=True), start=1):
        values = [str(cell).strip() if cell is not None else "" for cell in row]
        if "广告主名称" in values and "压价结果" in values:
            return row_idx, values
    raise ValueError(f"未在文件中找到表头行: {ws.title}")


def detect_yellow_headers(path: Path) -> list[str]:
    wb = load_workbook(path, read_only=False, data_only=False)
    ws = wb[wb.sheetnames[0]]
    header_row_idx, headers = find_header_row(ws)
    yellow_headers = []
    for col_idx, header in enumerate(headers, start=1):
        if not header:
            continue
        cell = ws.cell(header_row_idx, col_idx)
        if cell.fill and cell.fill.fill_type == "solid" and cell.fill.fgColor and cell.fill.fgColor.rgb == YELLOW_FILL_RGB:
            yellow_headers.append(header)
    return yellow_headers


def infer_feedback_month(period: str) -> int | None:
    match = re.search(r"(\d{1,2})月", period or "")
    if not match:
        return None
    month = int(match.group(1))
    return month if 1 <= month <= 12 else None


def resolve_feedback_template_path(explicit_path: str | None = None) -> Path | None:
    candidates = []
    if explicit_path:
        candidates.append(Path(explicit_path).expanduser().resolve())

    base_dir = Path(__file__).resolve().parent
    candidates.extend(
        [
            base_dir.parent / DEFAULT_FEEDBACK_TEMPLATE,
            base_dir / DEFAULT_FEEDBACK_TEMPLATE,
            Path.cwd() / DEFAULT_FEEDBACK_TEMPLATE,
        ]
    )

    for candidate in candidates:
        if candidate.exists():
            return candidate
    return None


def resolve_template_sheet_name(sheet_names: list[str], period: str) -> str | None:
    month = infer_feedback_month(period)
    if month:
        preferred = [
            f"{month}月影响毛利订单反馈",
            f"{month}月毛利影响严重订单反馈",
            f"{month}月毛利影响严重订单-反馈",
        ]
        for name in preferred:
            if name in sheet_names:
                return name
        fuzzy = [name for name in sheet_names if f"{month}月" in name and "反馈" in name]
        if fuzzy:
            return fuzzy[0]

    if DEFAULT_FEEDBACK_TEMPLATE_SHEET in sheet_names:
        return DEFAULT_FEEDBACK_TEMPLATE_SHEET

    feedback_sheets = [name for name in sheet_names if "反馈" in name]
    return feedback_sheets[0] if feedback_sheets else None


def get_template_sheet_xml_path(template_path: Path, sheet_name: str) -> str | None:
    workbook_ns = {
        "m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
        "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
        "rel": "http://schemas.openxmlformats.org/package/2006/relationships",
    }
    with zipfile.ZipFile(template_path) as archive:
        workbook_root = ET.fromstring(archive.read("xl/workbook.xml"))
        rels_root = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))

        target_rel_id = None
        for sheet in workbook_root.find("m:sheets", workbook_ns):
            if sheet.attrib.get("name") == sheet_name:
                target_rel_id = sheet.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
                break
        if not target_rel_id:
            return None

        for rel in rels_root:
            if rel.attrib.get("Id") == target_rel_id:
                target = rel.attrib.get("Target", "")
                if target:
                    return f"xl/{target}"
    return None


def extract_cell_style(cell) -> dict:
    return {
        "font": copy(cell.font),
        "fill": copy(cell.fill),
        "border": copy(cell.border),
        "alignment": copy(cell.alignment),
        "protection": copy(cell.protection),
        "number_format": cell.number_format,
    }


def apply_cell_style(cell, style: dict | None):
    if not style:
        return
    cell.font = copy(style["font"])
    cell.fill = copy(style["fill"])
    cell.border = copy(style["border"])
    cell.alignment = copy(style["alignment"])
    cell.protection = copy(style["protection"])
    cell.number_format = style["number_format"]


def load_template_layout(template_path: Path, sheet_name: str) -> dict:
    workbook = load_workbook(template_path, read_only=True, data_only=False)
    sheet = workbook[sheet_name]
    header_row = list(next(sheet.iter_rows(min_row=1, max_row=1)))
    sample_row = list(next(sheet.iter_rows(min_row=2, max_row=2)))

    column_width_by_index = {}
    default_column_width = 14.0
    default_row_height = 18.0
    freeze_panes = "A2"
    sheet_xml_path = get_template_sheet_xml_path(template_path, sheet_name)

    if sheet_xml_path:
        sheet_ns = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
        with zipfile.ZipFile(template_path) as archive:
            sheet_root = ET.fromstring(archive.read(sheet_xml_path))
        sheet_format = sheet_root.find("m:sheetFormatPr", sheet_ns)
        if sheet_format is not None:
            default_column_width = float(sheet_format.attrib.get("defaultColWidth", default_column_width))
            default_row_height = float(sheet_format.attrib.get("defaultRowHeight", default_row_height))

        pane = sheet_root.find("m:sheetViews/m:sheetView/m:pane", sheet_ns)
        if pane is not None and pane.attrib.get("topLeftCell"):
            freeze_panes = pane.attrib["topLeftCell"]

        cols = sheet_root.find("m:cols", sheet_ns)
        if cols is not None:
            for col in cols:
                width = float(col.attrib.get("width", default_column_width))
                for idx in range(int(col.attrib["min"]), int(col.attrib["max"]) + 1):
                    column_width_by_index[idx] = width

    header_styles = {}
    body_styles = {}
    column_widths = {}
    for col_idx, cell in enumerate(header_row, start=1):
        header = str(cell.value or "").strip()
        if not header:
            continue
        header_styles[header] = extract_cell_style(cell)
        column_widths[header] = column_width_by_index.get(col_idx, default_column_width)

    for col_idx, cell in enumerate(sample_row, start=1):
        header = str(header_row[col_idx - 1].value or "").strip()
        if not header:
            continue
        body_styles[header] = extract_cell_style(cell)

    workbook.close()
    return {
        "header_styles": header_styles,
        "body_styles": body_styles,
        "column_widths": column_widths,
        "freeze_panes": freeze_panes,
        "header_row_height": default_row_height,
    }


def load_source_highlight_layout(source_path: Path, headers_to_keep: list[str]) -> dict:
    workbook = load_workbook(source_path, read_only=False, data_only=False)
    sheet = workbook[workbook.sheetnames[0]]
    header_row_idx, headers = find_header_row(sheet)
    header_lookup = {str(value or "").strip(): idx for idx, value in enumerate(headers, start=1)}

    header_styles = {}
    body_styles = {}
    column_widths = {}
    base_body_style = None

    for header in headers_to_keep:
        col_idx = header_lookup.get(header)
        if not col_idx:
            continue
        header_cell = sheet.cell(header_row_idx, col_idx)
        header_styles[header] = extract_cell_style(header_cell)
        column_widths[header] = sheet.column_dimensions[get_column_letter(col_idx)].width or 12.0

        body_cell = sheet.cell(header_row_idx + 1, col_idx)
        style = extract_cell_style(body_cell)
        body_styles[header] = style
        if base_body_style is None:
            base_body_style = style

    workbook.close()
    return {
        "header_styles": header_styles,
        "body_styles": body_styles,
        "column_widths": column_widths,
        "freeze_panes": "A2",
        "header_row_height": 18.0,
        "body_base_style": base_body_style,
    }


def load_rows(path: Path, extra_headers=None):
    wb = load_workbook(path, read_only=False, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header_row_idx, headers = find_header_row(ws)
    index = {name: headers.index(label) for name, label in HEADER_REQUIRED.items()}
    extra_headers = extra_headers or []
    extra_index = {header: headers.index(header) for header in extra_headers if header in headers}

    parsed = []
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if not any(row):
            continue
        advertiser = str(row[index["advertiser"]] or "").strip()
        if not advertiser:
            continue
        price_result = str(row[index["price_result"]] or "").strip()
        rebate_sum = (
            to_float(row[index["rebate_tax_third"]])
            + to_float(row[index["rebate_tax_non_third"]])
            + to_float(row[index["rebate_tax_platform"]])
        )
        parsed.append(
            {
                "advertiser": advertiser,
                "purchase_order_no": str(row[index["purchase_order_no"]] or "").strip(),
                "is_sza": SZA_KEYWORD in advertiser,
                "spu_category": str(row[index["spu_category"]] or "").strip(),
                "exec_price_tax": to_float(row[index["exec_price_tax"]]),
                "gross_profit": to_float(row[index["gross_profit"]]),
                "gross_margin": to_percent_number(row[index["gross_margin"]]),
                "actual_income_no_tax": to_float(row[index["actual_income_no_tax"]]),
                "price_result": price_result,
                "rebate_sum": rebate_sum,
                "rebate_no_tax_estimate": to_float(row[index["rebate_no_tax_estimate"]]),
                "owner_name": extract_owner_name(row[index["owner_media"]]),
                "export_fields": {
                    header: ("" if row[col_idx] is None else str(row[col_idx]).strip())
                    for header, col_idx in extra_index.items()
                },
            }
        )
    return [row for row in parsed if EXCLUDED_SPU_KEYWORD not in row["spu_category"]]


def calc_core_metrics(rows, execute_rows, is_sza: bool):
    confirm_scope = [row for row in rows if row["is_sza"] == is_sza]
    execute_scope = [row for row in execute_rows if row["is_sza"] == is_sza]

    total_exec_price = sum(row["exec_price_tax"] for row in execute_scope)
    total_gross_profit = sum(row["gross_profit"] for row in confirm_scope)
    total_income_execute = sum(row["actual_income_no_tax"] for row in execute_scope)
    total_gross_profit_execute = sum(row["gross_profit"] for row in execute_scope)

    covered = [row for row in confirm_scope if row["price_result"] and row["rebate_sum"] != 0]
    success = [row for row in covered if row["price_result"] == "返点有提升"]

    execute_with_rebate = [row for row in execute_scope if row["rebate_sum"] != 0]
    rebate_no_tax_sum = sum(row["rebate_no_tax_estimate"] for row in execute_with_rebate)
    rebate_tax_sum = sum(row["rebate_sum"] for row in execute_with_rebate)

    return {
        "订单数量": str(len(confirm_scope)),
        "成交金额": fmt_money(total_exec_price),
        "预估毛利额": fmt_money(total_gross_profit),
        "预估毛利率": fmt_percent(total_gross_profit_execute, total_income_execute),
        "压价覆盖率": fmt_percent(len(covered), len(confirm_scope)),
        "压价成功率": fmt_percent(len(success), len(covered)),
        "压价提升毛利率": fmt_percent(rebate_no_tax_sum - rebate_tax_sum, total_income_execute),
    }


def calc_top5_non_sza(rows):
    scope = [row for row in rows if not row["is_sza"]]
    grouped = {}
    for row in scope:
        owner_name = row["owner_name"] or "未填写"
        bucket = grouped.setdefault(
            owner_name,
            {
                "owner_name": owner_name,
                "sample_advertiser": row["advertiser"],
                "exec_price_tax": 0.0,
                "gross_profit": 0.0,
                "actual_income_no_tax": 0.0,
            },
        )
        bucket["exec_price_tax"] += row["exec_price_tax"]
        bucket["gross_profit"] += row["gross_profit"]
        bucket["actual_income_no_tax"] += row["actual_income_no_tax"]

    top5 = sorted(grouped.values(), key=lambda item: item["exec_price_tax"], reverse=True)[:5]
    table = []
    for item in top5:
        table.append(
            [
                item["sample_advertiser"],
                fmt_money(item["exec_price_tax"]),
                fmt_money(item["gross_profit"]),
                fmt_percent(item["gross_profit"], item["actual_income_no_tax"]),
                "--",
                item["owner_name"],
            ]
        )
    return table


def calc_sza_spu_breakdown(confirm_rows, execute_rows):
    categories = sorted({row["spu_category"] for row in execute_rows if row["is_sza"] and row["spu_category"]})
    table = []
    for category in categories:
        confirm_scope = [row for row in confirm_rows if row["is_sza"] and row["spu_category"] == category]
        execute_scope = [row for row in execute_rows if row["is_sza"] and row["spu_category"] == category]
        covered = [row for row in confirm_scope if row["price_result"] and row["rebate_sum"] != 0]
        success = [row for row in covered if row["price_result"] == "返点有提升"]
        table.append(
            [
                category,
                str(len(execute_scope)),
                fmt_percent(len(covered), len(confirm_scope)),
                fmt_percent(len(success), len(covered)),
            ]
        )
    return sorted(table, key=lambda item: int(item[1]), reverse=True)


def build_raw_preview(rows, title, description, headers, links=None):
    preview_rows = []
    for row in rows[:10]:
        preview_rows.append([str(row.get(key, "")) for key in headers])
    return {
        "title": title,
        "description": description,
        "links": links or [],
        "headers": headers,
        "rows": preview_rows,
    }


def build_low_margin_orders(execute_rows):
    orders = []
    scoped_rows = sorted([r for r in execute_rows if r["gross_margin"] < LOW_MARGIN_THRESHOLD], key=lambda item: item["gross_margin"])
    for index, row in enumerate(scoped_rows):
        orders.append(
            {
                "id": row["purchase_order_no"] or f"{row['advertiser']}-{index}",
                "purchaseOrderNo": row["purchase_order_no"] or "--",
                "advertiser": row["advertiser"],
                "spuCategory": row["spu_category"],
                "execPriceTax": fmt_money(row["exec_price_tax"]),
                "grossMargin": f"{row['gross_margin']:.2f}%",
                "reason": "",
                "exportFields": row.get("export_fields", {}),
            }
        )
    return orders


def build_low_margin_feedback_excel(
    orders,
    yellow_headers,
    output_path: Path,
    period: str = "",
    template_path: Path | None = None,
    execute_source_path: Path | None = None,
):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "低毛利订单反馈表"
    sheet.freeze_panes = "A2"

    headers = yellow_headers + ["原因"]
    layout = None
    if execute_source_path and execute_source_path.exists():
        layout = load_source_highlight_layout(execute_source_path, yellow_headers)
    elif template_path and template_path.exists():
        template_wb = load_workbook(template_path, read_only=True, data_only=False)
        template_sheet_name = resolve_template_sheet_name(template_wb.sheetnames, period)
        template_wb.close()
        if template_sheet_name:
            layout = load_template_layout(template_path, template_sheet_name)
            sheet.title = template_sheet_name
            sheet.freeze_panes = layout["freeze_panes"]
            sheet.row_dimensions[1].height = layout["header_row_height"]

    yellow_fill = PatternFill(fill_type="solid", fgColor=YELLOW_FILL_RGB)
    header_base_style = None
    body_base_style = None
    if layout:
        header_base_style = next(iter(layout["header_styles"].values()), None)
        body_base_style = layout.get("body_base_style") or next(iter(layout["body_styles"].values()), None)

    for col_idx, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_idx, value=header)
        if layout:
            style = layout["header_styles"].get(header, header_base_style)
            if header == "原因" and style is None:
                cell.fill = yellow_fill
            else:
                apply_cell_style(cell, style)
            sheet.column_dimensions[get_column_letter(col_idx)].width = layout["column_widths"].get(header, 12.0 if header != "原因" else 16.0)
        else:
            cell.fill = yellow_fill

        if header == "原因" and layout:
            if header_base_style:
                apply_cell_style(cell, header_base_style)
            else:
                cell.fill = yellow_fill

    for row_idx, order in enumerate(orders, start=2):
        for col_idx, header in enumerate(yellow_headers, start=1):
            value = order.get("exportFields", {}).get(header, "")
            cell = sheet.cell(row=row_idx, column=col_idx, value=value)
            if layout:
                apply_cell_style(cell, layout["body_styles"].get(header, body_base_style))
        reason_cell = sheet.cell(row=row_idx, column=len(headers), value=order.get("reason", ""))
        if layout:
            apply_cell_style(reason_cell, body_base_style)

    option_sheet = workbook.create_sheet("原因选项")
    for idx, reason in enumerate(LOW_MARGIN_REASONS, start=1):
        option_sheet.cell(row=idx, column=1, value=reason)
    option_sheet.sheet_state = "hidden"

    formula = f"'原因选项'!$A$1:$A${len(LOW_MARGIN_REASONS)}"
    validation = DataValidation(type="list", formula1=formula, allow_blank=True)
    sheet.add_data_validation(validation)
    if len(orders) > 0:
        validation.add(f"{get_column_letter(len(headers))}2:{get_column_letter(len(headers))}{len(orders) + 1}")

    if not layout:
        for col_idx, header in enumerate(headers, start=1):
            if header == "原因":
                max_len = max(len(header), max((len(order.get("reason", "")) for order in orders), default=0), 8)
            else:
                max_len = max([len(str(header))] + [len(str(order.get("exportFields", {}).get(header, ""))) for order in orders])
            sheet.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 12), 36)

    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(len(orders) + 1, 1)}"
    workbook.save(output_path)


def summarize_low_margin_orders(orders):
    amount = sum(to_float(order["execPriceTax"]) for order in orders)
    return {
        "orderCount": str(len(orders)),
        "amount": fmt_money(amount),
        "filledCount": "0",
        "unfilledCount": str(len(orders)),
        "reasonRows": [["待填写", "0", "0.00", "先在低毛利订单反馈表中选择原因"]],
    }


def build_json(confirm_rows, execute_rows, confirm_path: Path, execute_path: Path, period: str, title: str, feedback_path: Path, yellow_headers: list[str]):
    sza_core = calc_core_metrics(confirm_rows, execute_rows, True)
    non_sza_core = calc_core_metrics(confirm_rows, execute_rows, False)
    spu_breakdown = calc_sza_spu_breakdown(confirm_rows, execute_rows)
    low_margin_orders = build_low_margin_orders(execute_rows)
    low_margin_summary = summarize_low_margin_orders(low_margin_orders)

    data = {
        "title": title,
        "period": period,
        "audience": "直属上级",
        "heroSummary": "经营数据按 SZA 与非SZA 拆分，成交金额切到执行订单口径；第二大项仅保留低毛利专项，并支持在低毛利订单反馈表中人工填写原因后汇总。",
        "metrics": [
            {"section": "core", "label": "SZA成交金额", "value": sza_core["成交金额"], "note": "来自执行订单口径。"},
            {"section": "core", "label": "非SZA成交金额", "value": non_sza_core["成交金额"], "note": "来自执行订单口径。"},
            {"section": "core", "label": "SZA预估毛利率", "value": sza_core["预估毛利率"], "note": "来自执行订单口径。"},
            {"section": "core", "label": "非SZA预估毛利率", "value": non_sza_core["预估毛利率"], "note": "来自执行订单口径。"},
            {"section": "margin", "label": "低毛利订单数", "value": low_margin_summary["orderCount"], "note": "执行订单中预估订单毛利率低于14%。"},
            {"section": "margin", "label": "低毛利成交金额", "value": low_margin_summary["amount"], "note": "来自执行价(含税)汇总。"},
            {"section": "margin", "label": "已填写原因", "value": low_margin_summary["filledCount"], "note": "人工填写后自动汇总。"},
            {"section": "risk", "label": "待协调事项", "value": "请维护", "note": "风险事项仍建议人工维护。"},
        ],
        "summaries": [
            {"title": "经营表现", "copy": "经营模块已经按 SZA 与非SZA 自动拆分，成交金额已切到执行订单口径。"},
            {"title": "专项推进", "copy": "第二大项已改成低毛利专项，低毛利订单来自执行订单中预估订单毛利率低于14%的数据。"},
            {"title": "事项管理", "copy": "风险与待协调事项建议继续由周会总结后人工录入，网页负责自动排版。"},
        ],
        "highlights": [
            {"title": "执行订单口径", "copy": f"成交金额、毛利率与提升毛利率来自：{execute_path.name}", "section": "core-section"},
            {"title": "汽车类目已过滤", "copy": f"计算前已删除 SPU类目 含“{EXCLUDED_SPU_KEYWORD}”的所有数据。", "section": "margin-section"},
            {"title": "重点客户TOP5", "copy": "非SZA核心指标汇总全部非华为订单，TOP5仅按执行订单成交金额单独摘录。", "section": "core-section"},
        ],
        "sections": {
            "core": {
                "quickLabel": "SZA / 非SZA 双分区",
                "sourceLinks": [{"label": "查看经营原表预览", "href": "#core-raw-source"}],
                "sza": {
                    "miniMetrics": [{"label": key, "value": value} for key, value in sza_core.items()],
                    "spuBreakdown": spu_breakdown,
                    "narratives": [{"title": "SZA经营结论", "copy": "SZA 依据广告主名称包含“华为”自动归类，并按 SPU类目 拆分压价结果。 "}],
                },
                "nonSza": {
                    "miniMetrics": [{"label": key, "value": value} for key, value in non_sza_core.items()],
                    "customers": calc_top5_non_sza(execute_rows),
                    "narratives": [{"title": "非SZA经营结论", "copy": "非SZA 上方核心指标汇总全部非华为订单，下面只额外摘录成交金额前5客户。"}],
                },
                "rawSource": build_raw_preview(
                    [
                        {
                            "来源": "确认订单",
                            "广告主名称": row["advertiser"],
                            "SPU类目": row["spu_category"],
                            "执行价(含税)": fmt_money(row["exec_price_tax"]),
                            "压价结果": row["price_result"],
                            "预估订单毛利额": fmt_money(row["gross_profit"]),
                        }
                        for row in confirm_rows[:10]
                    ],
                    "经营数据原表预览",
                    f"确认订单：{confirm_path.name}；执行订单：{execute_path.name}；已过滤 SPU类目 含“{EXCLUDED_SPU_KEYWORD}”的行。",
                    ["来源", "广告主名称", "SPU类目", "执行价(含税)", "压价结果", "预估订单毛利额"],
                    links=[
                        {"label": "打开确认订单原表", "href": to_file_href(confirm_path)},
                        {"label": "打开执行订单原表", "href": to_file_href(execute_path)},
                    ],
                ),
            },
            "margin": {
                "quickLabel": "低毛利订单识别 + 原因反馈",
                "sourceLinks": [
                    {"label": "打开可填写反馈表", "href": to_file_href(feedback_path)},
                    {"label": "打开执行订单原表", "href": to_file_href(execute_path)},
                ],
                "exportHeaders": yellow_headers,
                "miniMetrics": [
                    {"label": "低毛利订单数", "value": low_margin_summary["orderCount"]},
                    {"label": "低毛利成交金额", "value": low_margin_summary["amount"]},
                    {"label": "已填写原因", "value": low_margin_summary["filledCount"]},
                    {"label": "未填写原因", "value": low_margin_summary["unfilledCount"]},
                ],
                "reasons": low_margin_summary["reasonRows"],
                "orders": low_margin_orders,
                "feedbackFileHref": to_file_href(feedback_path),
                "narratives": [{"title": "专项结论", "copy": "低毛利订单反馈表支持人工补充原因，页面会根据填写结果自动刷新原因汇总。"}],
                "rawSource": build_raw_preview(
                    [
                        {
                            "采购订单编号": order["purchaseOrderNo"],
                            "广告主名称": order["advertiser"],
                            "SPU类目": order["spuCategory"],
                            "执行价(含税)": order["execPriceTax"],
                            "预估订单毛利率": order["grossMargin"],
                            "原因": order["reason"],
                        }
                        for order in low_margin_orders[:10]
                    ],
                    "低毛利订单原表预览",
                    f"执行订单来源：{execute_path.name}",
                    ["采购订单编号", "广告主名称", "SPU类目", "执行价(含税)", "预估订单毛利率", "原因"],
                    links=[
                        {"label": "打开批量填写反馈表", "href": to_file_href(feedback_path)},
                        {"label": "打开执行订单原表", "href": to_file_href(execute_path)},
                    ],
                ),
            },
            "risk": {
                "quickLabel": "周会总结自动排版",
                "sourceLinks": [{"label": "查看事项原表", "href": "#"}],
                "items": [
                    {
                        "title": "待填写事项1",
                        "category": "周会总结",
                        "progress": "请在导入后的 JSON 中直接维护本周进展。",
                        "risk": "请填写风险点。",
                        "support": "请填写需协调事项。",
                        "nextAction": "请填写下周动作。",
                        "owner": "行业二组",
                        "status": "watch",
                    }
                ],
            },
        },
    }
    return data


def main():
    parser = argparse.ArgumentParser(description="从采购订单明细表生成周报 JSON 数据。")
    parser.add_argument("--confirm-xlsx", required=True, help="确认订单 Excel 路径")
    parser.add_argument("--execute-xlsx", help="执行订单 Excel 路径；不传则默认与确认订单相同")
    parser.add_argument("--period", default="", help="周报周期，例如 2026年03月16日 - 2026年03月22日")
    parser.add_argument("--title", default="行业二组周报", help="周报标题")
    parser.add_argument("--output", default="weekly-report-data.generated.json", help="输出 JSON 路径")
    parser.add_argument("--feedback-template", default="", help="低毛利反馈导出模板路径，可选")
    args = parser.parse_args()

    confirm_path = Path(args.confirm_xlsx).expanduser().resolve()
    execute_path = Path(args.execute_xlsx).expanduser().resolve() if args.execute_xlsx else confirm_path
    output_path = Path(args.output).expanduser().resolve()
    feedback_path = output_path.with_name("低毛利订单反馈表.xlsx")
    feedback_template_path = resolve_feedback_template_path(args.feedback_template or None)

    confirm_rows = load_rows(confirm_path)
    yellow_headers = detect_yellow_headers(execute_path)
    execute_rows = load_rows(execute_path, extra_headers=yellow_headers)
    data = build_json(confirm_rows, execute_rows, confirm_path, execute_path, args.period, args.title, feedback_path, yellow_headers)
    build_low_margin_feedback_excel(
        data["sections"]["margin"]["orders"],
        yellow_headers,
        feedback_path,
        period=args.period,
        template_path=feedback_template_path,
        execute_source_path=execute_path,
    )

    output_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Created: {output_path}")
    print(f"Created: {feedback_path}")
    if feedback_template_path:
        print(f"Template: {feedback_template_path}")


if __name__ == "__main__":
    main()
